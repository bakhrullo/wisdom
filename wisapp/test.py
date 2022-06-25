# import openpyxl
# from django.contrib.auth.models import User
#
# from peewee import *
#
# conn = SqliteDatabase('db.sqlite3')
# cursor = conn.cursor()
#
#
# book = openpyxl.load_workbook('Книга1.xlsx')
# worksheet = book.active
# user = []
# passw = []
# for i in range(1, worksheet.max_row):
#     for col in worksheet.iter_cols(1, worksheet.max_column):
#         if len(str(col[i].value)) == 5:
#             passw.append(col[i].value)
#         else:
#             user.append(col[i].value)
#
# for i in range(1, len(user)+1):
#     User.create(username=passw[i], password=user[i])
# conn.close()

#
# import sqlite3
#
# conn = sqlite3.connect('mydatabase.db')
# cur = conn.cursor()
# users = {}
# curs = conn.cursor()
# curs.execute('''
# Create table if not exists USER(NAME varchar(100), PASSWORD varchar(100))
# ''')
# NAME = 'june'
# PASSWORD = 'salafan21'
# conn.executemany('INSERT INTO USER(NAME,PASSWORD) VALUES (? ,?);', [(NAME, PASSWORD)])
# conn.commit()


import openpyxl

book = openpyxl.load_workbook('PUPS.xlsx')
worksheet = book.active
p_code = []
p_grade = []
par_num = []
for i in range(1, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        if type(col[i].value) == int:
            if len(str(col[i].value)) == 5:
                p_code.append(col[i].value)
            else:
                p_grade.append(col[i].value)
        else:
            len(str(col[i].value)) == 5